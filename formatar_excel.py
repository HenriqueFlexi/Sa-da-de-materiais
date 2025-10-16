import os
import glob
from openpyxl import load_workbook # type: ignore
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side # type: ignore

def formatar_planilha(arquivo):
    if not os.path.exists(arquivo):
        print(f"Arquivo {arquivo} não encontrado.")
        return

    wb = load_workbook(arquivo)
    if "Saídas" not in wb.sheetnames:
        print(f"Aba 'Saídas' não encontrada no arquivo {arquivo}.")
        return

    ws = wb["Saídas"]

    fonte_negrito = Font(bold=True)
    preenchimento_cabecalho = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    alinhamento_centralizado = Alignment(horizontal="center", vertical="center")
    borda_fina = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Cabeçalho
    for cell in ws[1]:
        cell.font = fonte_negrito
        cell.fill = preenchimento_cabecalho
        cell.alignment = alinhamento_centralizado
        cell.border = borda_fina

    # Demais linhas
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = alinhamento_centralizado
            cell.border = borda_fina

    # Largura das colunas
    colunas = {
        "A": 20,  # Data/Hora
        "B": 20,  # Nome
        "C": 25,  # Projeto
        "D": 30,  # Material
        "E": 12,  # Quantidade
    }

    for col, largura in colunas.items():
        ws.column_dimensions[col].width = largura

    wb.save(arquivo)
    print(f"Formatação aplicada com sucesso ao arquivo {arquivo}!")

# Formatar todos os arquivos mensais
arquivos = glob.glob("saida_materiais_*.xlsx")
for arquivo in arquivos:
    formatar_planilha(arquivo)
